from openpyxl import load_workbook
from datetime import datetime, timedelta
from csv import reader, writer
from configparser import ConfigParser
from os import path

# ============================================== #
this_script_dir = path.dirname(path.realpath(__file__))

CONFIG_FILENAME = this_script_dir + '\\config.cfg'
config = ConfigParser()
config.read_file(open(CONFIG_FILENAME))

SOURCE_WORKBOOK_PATH = config.get('Paths Config', 'SOURCE_WORKBOOK_PATH')
RESULT_CSV_DIRECTORY = config.get('Paths Config', 'OUTPUT_CSV_DIRECTORY')
SETTINGS_CSV_FILENAME = this_script_dir + "\\Scrap_Settings.csv"

DATE_FORMAT = config.get('Common Config', 'OUTPUT_DATE_FORMAT')
DAYS_BEFORE_TODAY = int(config.get('Common Config', 'NUMBER_OF_DAYS_BEFORE_TODAY'))
DAYS_AFTER_TODAY = int(config.get('Common Config', 'NUMBER_OF_DAYS_AFTER_TODAY'))

DATA_SHEET_NAME = config.get('Workbook Settings', 'DATA_SHEET_NAME')
START_ROW = int(config.get('Workbook Settings', 'START_ROW'))
SHIFT_COLUMN = config.get('Workbook Settings', 'SHIFT_COLUMN')
DATE_COLUMN = config.get('Workbook Settings', 'DATE_COLUMN')
# ============================================== #


# Function reads machine names and their columns names fom CSV.
# Function returns dict with machine names and their columns name
# Function returns empty dict for filling with data
def read_machines_and_cells(CSV_filepath):
    result = dict()
    final_result = dict()

    with open(CSV_filepath, "r") as f:
        csv_reader = reader(f, delimiter=";")
        next(csv_reader)

        for line in csv_reader:
            column_names = line.pop(0)  # Get column names for future CSVs. Column names separated by ','
            key = line.pop(0)
            value = line
            result[key] = value
            final_result[key] = {'0': '', '1': [], '2': [], '3': [], '4': []}  # key = shift number, value = list with shift data 0 - for column header
            final_result[key]['0'] = column_names

    return result, final_result


# Function writes final prepared data into CSV files
def write_data_to_files(result_data):
    # csv_header = ['Date', 'Scrap Recycle', 'Scrap']

    for machine_name, shift_data in result_data.items():
        csv_header = shift_data['0'].split(',')  # Get list for CSV header
        del shift_data['0']  # Remove header info from shift data

        for shift_key in shift_data:
            csv_filename = "{}\\{}_Shift_{}_Report.csv".format(RESULT_CSV_DIRECTORY, machine_name, shift_key)

            with open(csv_filename, mode='w', newline='') as csv_file:
                csv_writer = writer(csv_file, delimiter=',')
                csv_writer.writerow(csv_header)

                for row in shift_data[shift_key]:
                    csv_writer.writerow(row)


# Function returns calculated date. It adds or subtract delta_days from|to today_date
def calculate_new_date(today_date, delta_days):
    new_date = today_date + timedelta(days=delta_days)

    current_year = today_date.year
    calculated_date_year = new_date.year

    if current_year > calculated_date_year:
        new_date = datetime.strptime('01/01/{} 00:00:00'.format(current_year), '%d/%m/%Y %H:%M:%S')

    if current_year < calculated_date_year:
        new_date = datetime.strptime('31/12/{} 00:00:00'.format(current_year), '%d/%m/%Y %H:%M:%S')

    return new_date


MACHINE_STATS, final_results = read_machines_and_cells(SETTINGS_CSV_FILENAME)

# Dates in string representation DD.MM.YYYY
today_date = datetime.today()
today_date_string = today_date.strftime(DATE_FORMAT)

start_date = calculate_new_date(today_date, -1 * DAYS_BEFORE_TODAY)
start_date_string = start_date.strftime(DATE_FORMAT)

end_date = calculate_new_date(today_date, DAYS_AFTER_TODAY)
end_date_string = end_date.strftime(DATE_FORMAT)

wb = load_workbook(SOURCE_WORKBOOK_PATH, data_only=True)
ws = wb[DATA_SHEET_NAME]

row_number = START_ROW
shift_date = today_date
is_show = False

while shift_date < end_date:
    shift_number = ws[SHIFT_COLUMN + str(row_number)].value

    if shift_number is None:
        row_number += 1
        continue

    shift_date = ws[DATE_COLUMN + str(row_number)].value

    if shift_date <= start_date:
        row_number += 1
        continue

    shift_date_string = shift_date.strftime(DATE_FORMAT)
    shift_number = str(shift_number)

    for machine_name, machine_column_names_with_data in MACHINE_STATS.items():
        data_to_add = list()
        data_to_add.append(shift_date_string)

        for column_name in machine_column_names_with_data:
            cell_name = column_name + str(row_number)
            cell_value = ws[cell_name].value

            if (cell_value == '#DIV/0!') or (cell_value == '#REF!'):
                cell_value = 0

            data_to_add.append(round(cell_value))
            # data_to_add.append(cell_value)

        # data_to_add.append(machine_name)
        if 'rec_waste' in machine_name:
            data_to_add[1] = round(data_to_add[1] / 10)  # Scrap recycle / 10
        final_results[machine_name][shift_number].append(data_to_add)

    row_number += 1

write_data_to_files(final_results)
